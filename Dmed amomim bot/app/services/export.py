from html import escape
from pathlib import Path
from tempfile import NamedTemporaryFile

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.models import Feedback, FeedbackStatus, FeedbackType
from app.domain.criteria import criterion_label, tag_label
from app.services.analytics import feedback_report_data
from app.services.feedback import average_rating


STATUS_LABELS = {
    FeedbackStatus.new: "Новый",
    FeedbackStatus.reviewed: "Изучено",
    FeedbackStatus.in_progress: "В работе",
    FeedbackStatus.closed: "Закрыто",
}

TYPE_LABELS = {
    FeedbackType.employee: "Сотрудник",
    FeedbackType.implementation: "Внедрение",
}


HEADERS = [
    "ID",
    "Тип",
    "Учреждение",
    "ФИО отправителя",
    "Телефон",
    "Сотрудник",
    "Дата",
    "Средняя оценка",
    "Оценки",
    "Метки",
    "Комментарий",
    "Telegram",
    "Статус",
]


def _type_label(feedback: Feedback) -> str:
    return TYPE_LABELS.get(feedback.feedback_type, feedback.feedback_type.value)


def _status_label(feedback: Feedback) -> str:
    return STATUS_LABELS.get(feedback.status, feedback.status.value)


def _ratings_text(feedback: Feedback) -> str:
    if not feedback.ratings:
        return "-"
    return "\n".join(
        f"{criterion_label(code, feedback.feedback_type.value)}: {value}/5"
        for code, value in feedback.ratings.items()
    )


def _tags_text(feedback: Feedback) -> str:
    return ", ".join(tag_label(code) for code in feedback.tags or []) or "-"


def _row(feedback: Feedback) -> list[str | float]:
    return [
        str(feedback.id),
        _type_label(feedback),
        feedback.institution.name,
        feedback.reviewer_full_name or "",
        feedback.reviewer_phone or "",
        feedback.employee.full_name if feedback.employee else "Не указан",
        feedback.created_at.strftime("%Y-%m-%d %H:%M"),
        average_rating(feedback.ratings),
        _ratings_text(feedback),
        _tags_text(feedback),
        feedback.comment or "",
        feedback.user.telegram_link or "",
        _status_label(feedback),
    ]


def _style_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="E8EEF8")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="1F2937")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 60)


async def export_feedback_xlsx(session: AsyncSession, institution_id: int | None = None) -> Path:
    data = await feedback_report_data(session, days=None, institution_id=institution_id)
    feedback_items = data["feedback_items"]
    summary = data["summary"]

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Сводка"
    summary_sheet.append(["Показатель", "Значение"])
    summary_rows = [
        ("Всего отзывов", summary["feedback_total"]),
        ("Сотрудников, от которых получен фидбек", summary["unique_reviewers_total"]),
        ("Средняя оценка", summary["average_rating"]),
        ("Отзывы по сотрудникам", summary["employee_feedback_total"]),
        ("Средняя оценка сотрудников", summary["employee_average_rating"]),
        ("Отзывы по внедрению", summary["implementation_feedback_total"]),
        ("Средняя оценка внедрения", summary["implementation_average_rating"]),
        ("С комментариями", summary["comments_total"]),
        ("Низкие оценки до 2.5", summary["low_score_total"]),
        ("Переходы по ссылкам", summary["link_visits_total"]),
    ]
    if data["institution"] is not None:
        summary_rows.insert(0, ("Учреждение", data["institution"].name))
    for label, value in summary_rows:
        summary_sheet.append([label, value])
    _style_sheet(summary_sheet)

    reviews_sheet = workbook.create_sheet("Отзывы")
    reviews_sheet.append(HEADERS)
    for item in feedback_items:
        reviews_sheet.append(_row(item))
    _style_sheet(reviews_sheet)

    institutions_sheet = workbook.create_sheet("Учреждения")
    institutions_sheet.append(["ID", "Учреждение", "Регион", "Отзывы", "Сотрудников ответило", "Средняя оценка"])
    for row in data["institutions"]:
        institutions_sheet.append(
            [
                row["id"],
                row["name"],
                row["region"] or "",
                row["feedback_total"],
                row["reviewers_total"],
                row["average_rating"],
            ]
        )
    _style_sheet(institutions_sheet)

    employees_sheet = workbook.create_sheet("Сотрудники")
    employees_sheet.append(["ID", "Сотрудник", "Учреждение", "Отзывы", "Средняя оценка"])
    for row in data["employees"]:
        employees_sheet.append(
            [row["id"], row["name"], row["institution"], row["feedback_total"], row["average_rating"]]
        )
    _style_sheet(employees_sheet)

    tmp = NamedTemporaryFile(prefix="dmed_feedback_", suffix=".xlsx", delete=False)
    tmp.close()
    workbook.save(tmp.name)
    return Path(tmp.name)


def _register_pdf_font() -> str:
    candidates = [
        Path("C:/Windows/Fonts/arial.ttf"),
        Path("C:/Windows/Fonts/calibri.ttf"),
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
    ]
    for path in candidates:
        if path.exists():
            pdfmetrics.registerFont(TTFont("DMEDFont", str(path)))
            return "DMEDFont"
    return "Helvetica"


def _paragraph(text: object, style: ParagraphStyle) -> Paragraph:
    return Paragraph(escape(str(text or "-")).replace("\n", "<br/>"), style)


async def export_feedback_pdf(session: AsyncSession, institution_id: int | None = None) -> Path:
    data = await feedback_report_data(session, days=None, institution_id=institution_id)
    summary = data["summary"]
    feedback_items = data["feedback_items"]

    tmp = NamedTemporaryFile(prefix="dmed_feedback_", suffix=".pdf", delete=False)
    tmp.close()

    font = _register_pdf_font()
    styles = getSampleStyleSheet()
    styles["Title"].fontName = font
    styles["Title"].fontSize = 18
    styles["Heading2"].fontName = font
    styles["Normal"].fontName = font
    styles["Normal"].fontSize = 9
    table_text = ParagraphStyle(
        "DMEDTableText",
        parent=styles["Normal"],
        fontName=font,
        fontSize=7,
        leading=9,
    )
    table_header = ParagraphStyle(
        "DMEDTableHeader",
        parent=table_text,
        fontName=font,
        fontSize=7,
        leading=9,
        textColor=colors.HexColor("#111827"),
    )

    doc = SimpleDocTemplate(
        tmp.name,
        pagesize=A4,
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
        title="Отчет DMED",
    )

    institution = data["institution"]
    title = "Отчет по отзывам DMED"
    if institution is not None:
        title += f": {institution.name}"

    story = [Paragraph(title, styles["Title"]), Spacer(1, 8)]
    summary_table = Table(
        [
            [_paragraph("Показатель", table_header), _paragraph("Значение", table_header)],
            [_paragraph("Всего отзывов", table_text), _paragraph(summary["feedback_total"], table_text)],
            [
                _paragraph("Сотрудников, от которых получен фидбек", table_text),
                _paragraph(summary["unique_reviewers_total"], table_text),
            ],
            [_paragraph("Средняя оценка", table_text), _paragraph(summary["average_rating"], table_text)],
            [_paragraph("Отзывы по сотрудникам", table_text), _paragraph(summary["employee_feedback_total"], table_text)],
            [_paragraph("Отзывы по внедрению", table_text), _paragraph(summary["implementation_feedback_total"], table_text)],
            [_paragraph("С комментариями", table_text), _paragraph(summary["comments_total"], table_text)],
            [_paragraph("Низкие оценки до 2.5", table_text), _paragraph(summary["low_score_total"], table_text)],
        ],
        colWidths=[105 * mm, 55 * mm],
    )
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8EEF8")),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    story.extend([summary_table, Spacer(1, 12), Paragraph("Последние отзывы", styles["Heading2"]), Spacer(1, 6)])

    table_rows = [
        [
            _paragraph("Дата", table_header),
            _paragraph("Учреждение", table_header),
            _paragraph("Тип", table_header),
            _paragraph("Отправитель", table_header),
            _paragraph("Телефон", table_header),
            _paragraph("Оценка", table_header),
            _paragraph("Комментарий", table_header),
        ]
    ]
    for item in feedback_items[:120]:
        subject = item.employee.full_name if item.employee else item.institution.name
        table_rows.append(
            [
                _paragraph(item.created_at.strftime("%Y-%m-%d %H:%M"), table_text),
                _paragraph(item.institution.name, table_text),
                _paragraph(f"{_type_label(item)}\n{subject}", table_text),
                _paragraph(item.reviewer_full_name or "-", table_text),
                _paragraph(item.reviewer_phone or "-", table_text),
                _paragraph(average_rating(item.ratings), table_text),
                _paragraph(item.comment or "-", table_text),
            ]
        )

    if len(table_rows) == 1:
        story.append(Paragraph("Отзывов пока нет.", styles["Normal"]))
    else:
        reviews_table = Table(
            table_rows,
            colWidths=[24 * mm, 34 * mm, 32 * mm, 31 * mm, 25 * mm, 16 * mm, 34 * mm],
            repeatRows=1,
        )
        reviews_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8EEF8")),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 3),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        story.append(reviews_table)

    doc.build(story)
    return Path(tmp.name)
